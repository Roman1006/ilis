# -*- coding: utf-8 -*-
# @Time : 2020-07-28 18:46
# @Author : Roman
# @FileName : Operation.py
# @Email : 13883575239@163.com
# @Software: PyCharm
import unittest
from selenium import webdriver
import HTMLTestRunner
import os
from openpyxl import load_workbook
from selenium import webdriver
import time
driver = webdriver.Chrome()
lib_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "../data"))
file_path = lib_path + "\\" + "测试哈哈.xlsx"  # excel的地址
class Operation():
    def login_out(self):
        driver = webdriver.Chrome()
        driver.find_element_by_link_text("退出系统").click()
        time.sleep(1)
        element = driver.find_element_by_css_selector("#layui-layer100001")
        element.find_element_by_class_name("layui-layer-btn0").click()
    def testopenBD(self):
        '''打开百度'''
        book = load_workbook(file_path)  # 默认可读写，若有需要可以指定write_only和read_only为True
        b = book["委托"]
        datadict = {} #创建一个字典变量
        maxC = b.max_column#获取最大列
        maxR = b.max_row#获取最大行
        for i in range(1, maxC + 1):#range默认从0开始，到后面参数的-1结束，而openpyxl都是从第一行第一列开始的，所以参数为1，maxC+1；意思就是遍历第一列到最后一列，
            datadict.setdefault(b.cell(1, i).value)#设置字典datadict的键值，从第一行第一列，到第一行最后一列，也就是第一行的标题设置为键值，i表示列
            # print(datadict)
        for i in range(2, maxR + 1):#除去第一行标题，从第二行遍历到最后一行
            for j in range(1, maxC + 1):#从第一列遍历到最后一列
                datadict[b.cell(1, j).value] = b.cell(i, j).value#在字典里设置键对应的值，j表示列
                # print(datadict)
            # if datadict["url"] & datadict["校验"] != None:
            self.driver.get(datadict["地址"])#从对应键里取出值地址对应的URL值，这里的键就是表格的第一行标题，所以EXCEL的标题要按照规定来写
            if self.driver.title == datadict["校验"]:#如果打开地址后，title值跟校验对应的值一致
                b.cell(i, maxC, '通过')#就将测试通过结果写入当前行的最后一列的单元格中，
            else:
                b.cell(i, maxC, '不通过')#就将测试不通过结果写入当前行的最后一列的单元格中，所以建立表格数据的时候，默认将结果写到最后一列即可
        book.save("mylogintest.xlsx")#最后记得关闭
