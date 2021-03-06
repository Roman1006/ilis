# -*- coding: utf-8 -*-
# @Time : 2020-12-28 17:27
# @Author : Roman
# @FileName : Operate_openpyxl.py
# @Email : 13883575239@163.com
# @Software: PyCharm
from iLIS import readconf
from openpyxl import load_workbook
localReadConfig = readconf.Read_conf()
class OperationExcel:
    """
    操作数据库
    """
    def __init__(self, excel_name, sheet_name=None):
        """
        构造方法，初始化实例属性
        :param excel_name: excel名称
        :param sheet_name: 表单名称
        """
        self.excel_name = excel_name
        self.sheet_name = sheet_name
        self.actual = OperationConfig("EXCEL", "actual").get_int_value()    # 从配置文件获取actual在第几列
        self.result = OperationConfig("EXCEL", "result").get_int_value()    # 从配置文件获取result在第几列

    def get_data(self):
        """
        获取测试数据
        :return: 测试数据
        """
        wb = load_workbook(self.excel_name)    # 创建workbook对象
        if self.sheet_name is None:    # 如果sheetname为空，默认激活第一个sheet，否则激活指定的sheet
            ws = wb.active
        else:
            ws =wb[self.sheet_name]
        header_data = tuple(ws.iter_rows(max_row=1, values_only=True))[0]   # 获取第一行的数据，转为元组
        other_data = tuple(ws.iter_rows(min_row=2, values_only=True))   # 获取其他行的数据，转为元组
        test_data = []
        for data in other_data:    # 将两个元组组合成字典，返回嵌套字典的列表
            data_dict = dict(zip(header_data, data))
            test_data.append(data_dict)
        return test_data

    def get_row_data(self, row):
        """
        获取某一行的数据
        :param row: 行数
        :return: 某一行的数据
        """
        return self.get_data()[row-1]

    def write_data(self, row, actual, result):
        """
        将数据写入excel
        :param row: 行号
        :param actual:实际结果
        :param result:执行状态
        :return:
        """
        wb = load_workbook(self.excel_name)
        if self.sheet_name is None:
            ws = wb.active
        else:
            ws = wb[self.sheet_name]
        if isinstance(row, int) and 2 <= row <= ws.max_row:
            ws.cell(row=row, column=self.actual, value=actual)
            ws.cell(row=row, column=self.result, value=result)
            wb.save(self.excel_name)
        else:
            print("请输入行号有误，应该大于1的正整数")